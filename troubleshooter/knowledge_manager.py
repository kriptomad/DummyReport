"""
🧠 Knowledge Base Manager - Gerenciamento Avançado de Base de Conhecimento
===========================================================================
Sistema de upload, categorização, matching e gestão de conhecimento de troubleshooting.
"""

import pandas as pd
import os
from typing import List, Dict, Tuple, Optional
from difflib import SequenceMatcher
from datetime import datetime
import re
from pathlib import Path
import json
import shutil
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity
import numpy as np

from troubleshooter import kb_ownership


class KnowledgeBaseManager:
    """Gerenciador centralizado de knowledge base"""

    def __init__(self, base_path: str = None):
        if base_path is None:
            base_path = Path(__file__).parent.parent

        self.base_path = Path(base_path)
        self.assets_path = self.base_path / 'assets'
        self.backup_path = self.assets_path / 'knowledge_base_backups'
        self.assets_path.mkdir(exist_ok=True)
        self.backup_path.mkdir(exist_ok=True)

        # Arquivo principal da knowledge base — aponta para o MESMO arquivo que
        # o motor de troubleshooting real usa (troubleshooter/loader.py,
        # assets/stepsdummy.xlsx, sheet "Main"). Isso garante que qualquer
        # upload/merge feito por esta UI realmente alimenta o Troubleshooter,
        # em vez de escrever num arquivo paralelo desconectado.
        self.kb_file = self.assets_path / 'stepsdummy.xlsx'
        self.kb_sheet = 'Main'
        self.metadata_file = self.assets_path / 'knowledge_base_metadata.json'

        # Colunas padrão esperadas
        self.required_columns = [
            'Categoria',
            'Mensagem de erro / padrão identificado',
            'Significado provável',
            'Precisa usar a Rate Card Lookup Query?',
            'Como validar',
            'Ação recomendada',
            'Responsável sugerido'
        ]

        # Mapeamento de aliases de colunas
        self.column_aliases = {
            'Mensagem de erro / padrão identificado': [
                'Error Message', 'Error Pattern', 'Error', 'Message',
                'Erro', 'Mensagem', 'Pattern', 'Padrão', 'Mensagem de erro',
                'Error_Message', 'ErrorMessage', 'error_pattern'
            ],
            'Categoria': [
                'Category', 'Type', 'Error Type', 'Tipo', 'Cat',
                'Error_Category', 'ErrorCategory', 'error_type'
            ],
            'Significado provável': [
                'Meaning', 'Description', 'Significado', 'Descrição',
                'Probable_Meaning', 'ProbableMeaning', 'meaning'
            ],
            'Como validar': [
                'Validation', 'How to Validate', 'Validação', 'How_to_Check',
                'HowToValidate', 'validation', 'check'
            ],
            'Ação recomendada': [
                'Action', 'Recommended Action', 'Solution', 'Ação', 'Solução',
                'Recommended_Action', 'RecommendedAction', 'action', 'solution'
            ],
            'Responsável sugerido': [
                'Owner', 'Responsible', 'Team', 'Responsável', 'Responsavel',
                'Suggested_Owner', 'SuggestedOwner', 'owner', 'responsible'
            ],
            'Precisa usar a Rate Card Lookup Query?': [
                'Tariff Query', 'Needs Tariff Query', 'Usa Tariff', 'Needs_Tariff',
                'NeedsTariff', 'needs_tariff', 'tariff_query'
            ]
        }

        # Cache da knowledge base
        self._kb_cache = None
        self._kb_cache_time = None
        self._kb_cache_file_mtime = None

        # TF-IDF para matching inteligente
        self._tfidf_vectorizer = None
        self._tfidf_matrix = None

    def standardize_columns(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Padroniza nomes de colunas usando aliases

        Args:
            df: DataFrame com colunas variadas

        Returns:
            DataFrame com colunas padronizadas
        """
        # Cria mapeamento reverso
        reverse_map = {}
        for standard_name, aliases in self.column_aliases.items():
            for alias in aliases:
                reverse_map[alias.lower().strip()] = standard_name

        # Renomeia colunas
        renamed_columns = {}
        for col in df.columns:
            col_lower = str(col).lower().strip()
            if col_lower in reverse_map:
                renamed_columns[col] = reverse_map[col_lower]

        df = df.rename(columns=renamed_columns)

        # Adiciona colunas faltantes com valores vazios
        for col in self.required_columns:
            if col not in df.columns:
                df[col] = ''

        # Remove linhas completamente vazias
        df = df.dropna(how='all')

        # Retorna apenas colunas padronizadas
        return df[self.required_columns]

    def get_current_version(self) -> str:
        """
        Retorna versão atual da KB

        Returns:
            String no formato "X.Y.Z"
        """
        if self.metadata_file.exists():
            try:
                with open(self.metadata_file, 'r', encoding='utf-8') as f:
                    metadata = json.load(f)
                return metadata.get('version', '1.0.0')
            except:
                return '1.0.0'

        return '1.0.0'

    def increment_version(self, version_type: str = 'patch') -> str:
        """
        Incrementa versão

        Args:
            version_type: 'major', 'minor', 'patch'

        Returns:
            Nova versão no formato "X.Y.Z"
        """
        current = self.get_current_version()
        major, minor, patch = map(int, current.split('.'))

        if version_type == 'major':
            return f"{major + 1}.0.0"
        elif version_type == 'minor':
            return f"{major}.{minor + 1}.0"
        else:  # patch
            return f"{major}.{minor}.{patch + 1}"

    def create_backup(self) -> Optional[Path]:
        """
        Cria backup da versão atual antes de update

        Returns:
            Path do arquivo de backup ou None se não existe KB atual
        """
        if not self.kb_file.exists():
            return None

        # Nome do backup: YYYY-MM-DD_HH-MM-SS_vX.X.X.xlsx
        timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
        version = self.get_current_version()
        backup_name = f"kb_{timestamp}_v{version}.xlsx"
        backup_file = self.backup_path / backup_name

        # Copia
        shutil.copy2(self.kb_file, backup_file)

        # Limita número de backups (mantém últimos 10)
        backups = sorted(self.backup_path.glob('kb_*.xlsx'), key=lambda p: p.stat().st_mtime, reverse=True)
        for old_backup in backups[10:]:
            old_backup.unlink()

        return backup_file

    def save_metadata(self, version: str, description: str = '', stats: dict = None):
        """
        Salva metadata da versão atual

        Args:
            version: Versão atual
            description: Descrição das mudanças
            stats: Estatísticas adicionais
        """
        metadata = {
            'version': version,
            'last_updated': datetime.now().isoformat(),
            'description': description,
            'stats': stats or {}
        }

        with open(self.metadata_file, 'w', encoding='utf-8') as f:
            json.dump(metadata, f, indent=2, ensure_ascii=False)

    def load_knowledge_base(self, force_reload: bool = False) -> pd.DataFrame:
        """
        Carrega a knowledge base com cache inteligente

        Args:
            force_reload: Força recarregar do disco

        Returns:
            DataFrame com a knowledge base
        """
        current_mtime = os.path.getmtime(self.kb_file) if self.kb_file.exists() else None

        # Verifica se tem cache válido (< 5 minutos) e se o arquivo não mudou.
        if not force_reload and self._kb_cache is not None:
            if (
                self._kb_cache_time
                and current_mtime is not None
                and self._kb_cache_file_mtime == current_mtime
                and (datetime.now() - self._kb_cache_time).seconds < 300
            ):
                return self._kb_cache.copy()

        # Carrega do disco
        if self.kb_file.exists():
            df = pd.read_excel(self.kb_file, sheet_name=self.kb_sheet)

            # Padroniza colunas
            df = self.standardize_columns(df)

            # Remove linhas sem mensagem de erro
            df = df[df['Mensagem de erro / padrão identificado'].astype(str).str.strip() != ''].copy()

            # Atualiza cache
            self._kb_cache = df.copy()
            self._kb_cache_time = datetime.now()
            self._kb_cache_file_mtime = current_mtime

            # Atualiza TF-IDF
            self._update_tfidf(df)

            return df
        else:
            # Retorna DataFrame vazio com colunas corretas
            return pd.DataFrame(columns=self.required_columns)

    def upload_and_merge(self, uploaded_file_path: str, file_type: str = 'auto',
                        version_type: str = 'patch', description: str = '',
                        cws: Optional[str] = None) -> Dict:
        """
        Faz upload de novo arquivo e mescla com knowledge base existente
        (assets/stepsdummy.xlsx, sheet "Main" — o MESMO arquivo usado pelo
        motor de troubleshooting real). Preserva as demais sheets (Logic,
        Steps, Type of errors). Após o merge, recarrega o motor de
        troubleshooting para que as mudanças tenham efeito imediato.

        Args:
            uploaded_file_path: Caminho do arquivo enviado
            file_type: 'excel', 'csv', 'docx' ou 'auto' para detectar
            version_type: 'major', 'minor', 'patch'
            description: Descrição das mudanças
            cws: CWS do usuário que está enviando o arquivo (para
                 atribuição de ownership das entradas novas/atualizadas)

        Returns:
            Estatísticas do merge
        """
        file_path = Path(uploaded_file_path)

        # Detecta tipo automaticamente
        if file_type == 'auto':
            ext = file_path.suffix.lower()
            if ext in ['.xlsx', '.xls']:
                file_type = 'excel'
            elif ext == '.csv':
                file_type = 'csv'
            elif ext == '.docx':
                file_type = 'docx'
            else:
                raise ValueError(f"Formato de arquivo não suportado: {ext}")

        # Carrega novo arquivo
        if file_type == 'excel':
            df_new = pd.read_excel(uploaded_file_path)
        elif file_type == 'csv':
            df_new = pd.read_csv(uploaded_file_path)
        elif file_type == 'docx':
            # Para DOCX, precisa de processamento especial (não implementado ainda)
            raise NotImplementedError("Suporte para DOCX será implementado em breve")
        else:
            raise ValueError(f"Tipo de arquivo não suportado: {file_type}")

        # Padroniza colunas
        df_new = self.standardize_columns(df_new)

        # Carrega KB atual
        df_current = self.load_knowledge_base()
        pattern_col = 'Mensagem de erro / padrão identificado'
        existing_patterns = set(df_current[pattern_col].str.strip().str.lower()) if not df_current.empty else set()
        new_patterns = set(df_new[pattern_col].str.strip().str.lower())

        # Estatísticas antes do merge
        stats_before = {
            'total': len(df_current),
            'categories': df_current['Categoria'].nunique() if not df_current.empty else 0
        }

        # Cria backup (copia o arquivo INTEIRO, todas as sheets)
        backup_file = self.create_backup()

        # Merge: remove duplicatas baseado em "Mensagem de erro"
        df_merged = pd.concat([df_current, df_new], ignore_index=True)

        # Remove duplicatas mantendo a última entrada (mais recente)
        df_merged = df_merged.drop_duplicates(
            subset=[pattern_col],
            keep='last'
        )

        # Estatísticas depois do merge
        stats_after = {
            'total': len(df_merged),
            'categories': df_merged['Categoria'].nunique(),
            'new_entries': len(df_new),
            'duplicates_removed': len(df_current) + len(df_new) - len(df_merged)
        }

        # Salva nova versão (apenas a sheet "Main", preservando as demais)
        new_version = self.increment_version(version_type)
        self._write_main_sheet(df_merged)

        # Stamp ownership: patterns that already existed get "updated",
        # brand-new patterns get "created" — attributed to the uploader's CWS.
        for pattern in new_patterns:
            if pattern in existing_patterns:
                kb_ownership.stamp_updated(pattern, cws or kb_ownership.SYSTEM_OWNER)
            else:
                kb_ownership.stamp_created(pattern, cws or kb_ownership.SYSTEM_OWNER)

        # Salva metadata
        self.save_metadata(
            version=new_version,
            description=description or f"Merged {len(df_new)} new entries",
            stats={'before': stats_before, 'after': stats_after, 'backup': str(backup_file) if backup_file else None}
        )

        # Invalida cache local e recarrega o motor de troubleshooting real
        self._kb_cache = None
        self._kb_cache_time = None
        self._kb_cache_file_mtime = None
        from troubleshooter.loader import load_all as _reload_troubleshoot_engine
        _reload_troubleshoot_engine(force_reload=True)

        return {
            'success': True,
            'version': new_version,
            'stats': stats_after,
            'backup_file': str(backup_file) if backup_file else None
        }

    def _write_main_sheet(self, df_main: pd.DataFrame) -> None:
        """Writes df_main back to the 'Main' sheet of stepsdummy.xlsx, preserving all other sheets."""
        from openpyxl import load_workbook
        from openpyxl.utils.dataframe import dataframe_to_rows

        wb = load_workbook(self.kb_file)
        if self.kb_sheet in wb.sheetnames:
            idx = wb.sheetnames.index(self.kb_sheet)
            del wb[self.kb_sheet]
        else:
            idx = 0
        ws = wb.create_sheet(self.kb_sheet, idx)

        for row in dataframe_to_rows(df_main, index=False, header=True):
            ws.append(row)

        wb.save(self.kb_file)

    def categorize_errors(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Categoriza erros automaticamente baseado em padrões

        Args:
            df: DataFrame com erros

        Returns:
            DataFrame com categoria atribuída
        """
        # Regras de categorização baseadas em keywords
        category_rules = {
            'Master Data': ['master data', 'location', 'shpg_loc', 'orig', 'dest', 'origin', 'destination'],
            'Rate/Tariff': ['rate', 'tariff', 'charge', 'tff', 'chrg', 'pricing'],
            'Equipment': ['equipment', 'container', 'eqmt', 'equip'],
            'Service': ['service', 'srvc', 'carrier', 'carr'],
            'Validation': ['validation', 'invalid', 'not found', 'missing'],
            'System': ['system', 'error', 'failed', 'timeout']
        }

        def assign_category(error_msg: str) -> str:
            if pd.isna(error_msg):
                return 'Uncategorized'

            error_lower = str(error_msg).lower()

            for category, keywords in category_rules.items():
                if any(kw in error_lower for kw in keywords):
                    return category

            return 'Uncategorized'

        # Atribui categoria apenas se estiver vazia
        if 'Categoria' in df.columns:
            df['Categoria'] = df.apply(
                lambda row: assign_category(row['Mensagem de erro / padrão identificado'])
                if pd.isna(row['Categoria']) or str(row['Categoria']).strip() == ''
                else row['Categoria'],
                axis=1
            )

        return df

    def _update_tfidf(self, df: pd.DataFrame):
        """
        Atualiza matriz TF-IDF para matching inteligente

        Args:
            df: DataFrame da knowledge base
        """
        if df.empty:
            self._tfidf_vectorizer = None
            self._tfidf_matrix = None
            return

        # Extrai textos de erro
        error_patterns = df['Mensagem de erro / padrão identificado'].fillna('').astype(str).tolist()

        # Cria vetorizador TF-IDF
        self._tfidf_vectorizer = TfidfVectorizer(
            max_features=1000,
            ngram_range=(1, 3),
            stop_words=None,
            lowercase=True
        )

        # Gera matriz
        self._tfidf_matrix = self._tfidf_vectorizer.fit_transform(error_patterns)

    def find_similar_errors(self, error_msg: str, top_k: int = 5, threshold: float = 0.3) -> List[Dict]:
        """
        Encontra erros similares usando TF-IDF + cosine similarity

        Args:
            error_msg: Mensagem de erro para buscar
            top_k: Número de resultados a retornar
            threshold: Score mínimo de similaridade (0-1)

        Returns:
            Lista de dicts com matches e scores
        """
        if self._tfidf_vectorizer is None or self._tfidf_matrix is None:
            # Fallback para método simples
            return []

        # Vetoriza a query
        query_vector = self._tfidf_vectorizer.transform([error_msg])

        # Calcula similaridade
        similarities = cosine_similarity(query_vector, self._tfidf_matrix).flatten()

        # Pega top-k acima do threshold
        top_indices = np.argsort(similarities)[::-1][:top_k]

        kb = self.load_knowledge_base()
        results = []

        for idx in top_indices:
            score = float(similarities[idx])
            if score >= threshold:
                row = kb.iloc[idx].to_dict()
                row['_match_score'] = round(score, 3)
                results.append(row)

        return results

    def get_statistics(self) -> Dict:
        """
        Retorna estatísticas da knowledge base

        Returns:
            Dicionário com estatísticas
        """
        df = self.load_knowledge_base()

        if df.empty:
            return {
                'total_entries': 0,
                'categories': 0,
                'version': self.get_current_version()
            }

        return {
            'total_entries': len(df),
            'categories': df['Categoria'].nunique(),
            'category_breakdown': df['Categoria'].value_counts().to_dict(),
            'needs_tariff': df['Precisa usar a Rate Card Lookup Query?'].value_counts().to_dict(),
            'version': self.get_current_version(),
            'last_updated': self._kb_cache_time.isoformat() if self._kb_cache_time else None
        }

    def export_to_excel(self, output_path: str, include_stats: bool = True):
        """
        Exporta knowledge base para Excel com múltiplas abas

        Args:
            output_path: Caminho do arquivo de saída
            include_stats: Se deve incluir aba de estatísticas
        """
        df = self.load_knowledge_base()

        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Aba principal
            df.to_excel(writer, sheet_name='All Errors', index=False)

            # Aba por categoria
            if 'Categoria' in df.columns:
                for category in df['Categoria'].unique():
                    if pd.notna(category) and str(category).strip():
                        df_cat = df[df['Categoria'] == category]
                        sheet_name = str(category)[:31]  # Excel limit
                        df_cat.to_excel(writer, sheet_name=sheet_name, index=False)

            # Aba de estatísticas
            if include_stats:
                stats = self.get_statistics()
                df_stats = pd.DataFrame([stats])
                df_stats.to_excel(writer, sheet_name='Statistics', index=False)

        return output_path

    def _import_excel(self, file_path: Path) -> pd.DataFrame:
        """Importa dados de Excel"""
        # Tenta ler primeira planilha
        df = pd.read_excel(file_path, sheet_name=0)
        return self._standardize_dataframe(df)

    def _import_csv(self, file_path: Path) -> pd.DataFrame:
        """Importa dados de CSV"""
        # Tenta com diferentes encodings
        for encoding in ['utf-8', 'latin-1', 'cp1252']:
            try:
                df = pd.read_csv(file_path, encoding=encoding)
                return self._standardize_dataframe(df)
            except UnicodeDecodeError:
                continue

        raise ValueError("Não foi possível decodificar o arquivo CSV")

    def _import_docx(self, file_path: Path) -> pd.DataFrame:
        """Importa dados de Word (tabelas)"""
        from docx import Document

        doc = Document(file_path)
        all_data = []

        for table in doc.tables:
            # Primeira linha são headers
            headers = [cell.text.strip() for cell in table.rows[0].cells]

            # Demais linhas são dados
            for row in table.rows[1:]:
                row_data = [cell.text.strip() for cell in row.cells]
                if len(row_data) == len(headers):
                    all_data.append(dict(zip(headers, row_data)))

        if not all_data:
            raise ValueError("Nenhuma tabela encontrada no documento Word")

        df = pd.DataFrame(all_data)
        return self._standardize_dataframe(df)

    def _standardize_dataframe(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Padroniza DataFrame para formato da knowledge base
        """
        # Mapeamento de colunas comuns
        column_mapping = {
            # Variações de Categoria
            'Category': 'Categoria',
            'Error Category': 'Categoria',
            'Type': 'Categoria',
            'Error Type': 'Categoria',

            # Variações de Mensagem de Erro
            'Error Message': 'Mensagem de erro / padrão identificado',
            'Error Pattern': 'Mensagem de erro / padrão identificado',
            'Error': 'Mensagem de erro / padrão identificado',
            'Message': 'Mensagem de erro / padrão identificado',
            'Pattern': 'Mensagem de erro / padrão identificado',

            # Variações de Significado
            'Meaning': 'Significado provável',
            'Description': 'Significado provável',
            'Root Cause': 'Significado provável',
            'Cause': 'Significado provável',
            'Probable Meaning': 'Significado provável',

            # Variações de Tariff Query
            'Tariff Query': 'Precisa usar a Rate Card Lookup Query?',
            'Needs Tariff Query': 'Precisa usar a Rate Card Lookup Query?',
            'Use Tariff Query': 'Precisa usar a Rate Card Lookup Query?',

            # Variações de Validação
            'Validation': 'Como validar',
            'How to Validate': 'Como validar',
            'Validation Steps': 'Como validar',

            # Variações de Ação
            'Action': 'Ação recomendada',
            'Recommended Action': 'Ação recomendada',
            'Action Required': 'Ação recomendada',
            'Resolution': 'Ação recomendada',
            'Fix': 'Ação recomendada',

            # Variações de Responsável
            'Owner': 'Responsável sugerido',
            'Responsible': 'Responsável sugerido',
            'Team': 'Responsável sugerido',
            'Assigned To': 'Responsável sugerido',
        }

        # Renomeia colunas
        df = df.rename(columns=column_mapping)

        # Garante que todas as colunas necessárias existem
        for col in self.required_columns:
            if col not in df.columns:
                df[col] = ''

        # Seleciona apenas as colunas necessárias
        df = df[self.required_columns]

        # Limpa dados
        df = df.fillna('')

        # Remove linhas completamente vazias
        df = df[df['Mensagem de erro / padrão identificado'].str.strip() != '']

        return df

    def smart_match_errors(self, error_messages: List[str], threshold: float = 0.45) -> List[Dict]:
        """
        Faz matching inteligente de múltiplas mensagens de erro

        Args:
            error_messages: Lista de mensagens de erro
            threshold: Limite de similaridade (0.0 a 1.0)

        Returns:
            Lista de dicionários com matches para cada erro
        """
        kb = self.load_knowledge_base()
        results = []

        for error_msg in error_messages:
            matches = self._match_single_error(error_msg, kb, threshold)
            results.append({
                'error_message': error_msg,
                'matches': matches,
                'best_match': matches[0] if matches else None
            })

        return results

    def _match_single_error(self, error_msg: str, kb: pd.DataFrame, threshold: float) -> List[Dict]:
        """Match de um único erro"""
        matches = []
        error_lower = str(error_msg).lower().strip()

        if not error_lower:
            return matches

        for idx, row in kb.iterrows():
            pattern = str(row['Mensagem de erro / padrão identificado']).lower().strip()

            if not pattern:
                continue

            # Calcula similaridade usando SequenceMatcher
            similarity = SequenceMatcher(None, error_lower, pattern).ratio()

            # Boost para palavras-chave importantes
            keywords = self._extract_keywords(error_lower)
            pattern_keywords = self._extract_keywords(pattern)

            common_keywords = keywords & pattern_keywords
            keyword_boost = len(common_keywords) * 0.05  # 5% por keyword em comum

            # Score final
            final_score = min(similarity + keyword_boost, 1.0)

            if final_score >= threshold:
                matches.append({
                    'score': final_score,
                    'similarity': similarity,
                    'keyword_boost': keyword_boost,
                    'categoria': row['Categoria'],
                    'padrao': row['Mensagem de erro / padrão identificado'],
                    'significado': row['Significado provável'],
                    'precisa_tariff_query': row['Precisa usar a Rate Card Lookup Query?'],
                    'como_validar': row['Como validar'],
                    'acao_recomendada': row['Ação recomendada'],
                    'responsavel': row['Responsável sugerido']
                })

        # Ordena por score decrescente
        matches.sort(key=lambda x: x['score'], reverse=True)

        # Retorna top 5
        return matches[:5]

    def _extract_keywords(self, text: str) -> set:
        """Extrai keywords importantes de um texto"""
        # Remove pontuação e números
        text = re.sub(r'[^\w\s]', ' ', text)
        text = re.sub(r'\d+', '', text)

        # Palavras importantes (não remove stop words básicas)
        words = set(text.split())

        # Remove palavras muito curtas
        words = {w for w in words if len(w) > 2}

        return words

    def categorize_error_auto(self, error_message: str) -> str:
        """
        Categoriza automaticamente um erro baseado em padrões

        Args:
            error_message: Mensagem de erro

        Returns:
            Categoria identificada
        """
        msg_lower = error_message.lower()

        # Regras de categorização
        if 'missing' in msg_lower or 'not found' in msg_lower:
            if 'tariff' in msg_lower or 'rate' in msg_lower:
                return 'Tariff/Rate - Missing Data'
            elif 'lane' in msg_lower or 'route' in msg_lower:
                return 'Lane/Route - Missing Data'
            else:
                return 'General - Missing Data'

        elif 'invalid' in msg_lower:
            if 'date' in msg_lower or 'expired' in msg_lower:
                return 'Date/Validity Issue'
            elif 'equipment' in msg_lower or 'equip' in msg_lower:
                return 'Equipment Issue'
            else:
                return 'General - Invalid Data'

        elif 'duplicate' in msg_lower:
            return 'Duplicate Data'

        elif 'expired' in msg_lower or 'date' in msg_lower:
            return 'Date/Expiration Issue'

        elif 'tariff' in msg_lower or 'rate' in msg_lower:
            return 'Tariff/Rate Issue'

        elif 'lane' in msg_lower or 'origin' in msg_lower or 'destination' in msg_lower:
            return 'Lane/Routing Issue'

        elif 'equipment' in msg_lower or 'equip' in msg_lower:
            return 'Equipment Issue'

        elif 'service' in msg_lower:
            return 'Service Issue'

        elif 'carrier' in msg_lower:
            return 'Carrier Issue'

        else:
            return 'General/Other'

    def export_to_csv(self, output_path: str = None) -> str:
        """Exporta knowledge base para CSV"""
        if output_path is None:
            output_path = self.assets_path / f'knowledge_base_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'

        kb = self.load_knowledge_base()
        kb.to_csv(output_path, index=False, encoding='utf-8-sig')

        return str(output_path)
