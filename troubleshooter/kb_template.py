"""
troubleshooter/kb_template.py
==============================
Generates a downloadable "Knowledge Feeder File" — a predefined Excel layout
that matches the exact schema the app's troubleshooting engine expects
(troubleshooter/loader.py column names), so that when someone fills it in
with many new errors/fixes and uploads it via the Knowledge Base tab's
"Upload & Merge", the app can parse it with zero ambiguity.

Includes:
  - "KB Feeder" sheet: the actual columns to fill in, with one example row
    (clearly marked, safe to delete) and dropdown data validation for the
    Tariff-Query and Category fields.
  - "Instructions" sheet: plain-language explanation of every column and
    how the merge/matching logic will use it.
"""
import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

from troubleshooter.loader import (
    COL_CATEGORY,
    COL_ERROR_PATTERN,
    COL_MEANING,
    COL_NEEDS_TARIFF,
    COL_HOW_TO_CHECK,
    COL_ACTION,
    COL_RESPONSIBLE,
    get_error_types_df,
)

FEEDER_COLUMNS = [
    COL_CATEGORY,
    COL_ERROR_PATTERN,
    COL_MEANING,
    COL_NEEDS_TARIFF,
    COL_HOW_TO_CHECK,
    COL_ACTION,
    COL_RESPONSIBLE,
]

EXAMPLE_ROW = [
    "Missing Rate / Lane",
    "EXAMPLE — replace with the real error message text (delete this row before uploading)",
    "The Routing system could not find a valid rate/lane for this Origin+Destination+Equipment combination.",
    "Sim",
    "Check the Rate Card Lookup Query for this lane; confirm rate exists and is active.",
    "If rate doesn't exist, escalate to Procurement with lane details. If it exists, re-run the shipment.",
    "Procurement Team",
]

INSTRUCTIONS = [
    ("Column", "What to put here"),
    (COL_CATEGORY, "A short category label (e.g. 'Missing Rate / Lane', 'Equipment Configuration'). "
                   "If you're not sure, leave it blank — the app will try to auto-classify it."),
    (COL_ERROR_PATTERN, "REQUIRED. The exact (or representative) error message text as it appears in "
                        "ERR_MSG. This is the key the app matches against — be as precise as possible."),
    (COL_MEANING, "One sentence: what does this error actually mean / what's the likely root cause?"),
    (COL_NEEDS_TARIFF, "Type 'Sim' or 'Não' (Yes/No): does resolving this error usually require checking "
                       "the Rate Card Lookup Query?"),
    (COL_HOW_TO_CHECK, "Short steps: how does an analyst confirm this is really the issue?"),
    (COL_ACTION, "REQUIRED. The recommended fix / resolution steps."),
    (COL_RESPONSIBLE, "Who should own/resolve this type of error (team or role name)."),
    ("", ""),
    ("Notes:", ""),
    ("1.", "Do not rename or reorder the columns — the app matches them by exact header name."),
    ("2.", "One row = one error pattern. Add as many rows as you need."),
    ("3.", "Delete the EXAMPLE row before uploading — it is just a reference."),
    ("4.", "Upload this file in the app: Knowledge Base tab -> Upload & Merge -> choose this .xlsx."),
    ("5.", "Existing entries with the exact same error message text will be UPDATED (not duplicated); "
           "brand-new error messages will be ADDED as new knowledge base entries."),
    ("6.", "Whoever uploads the file is recorded as the owner/author of any new or updated entries."),
]


def generate_feeder_template() -> bytes:
    """Builds the 'Knowledge Feeder File' template and returns it as .xlsx bytes."""
    wb = Workbook()

    # ── Sheet 1: KB Feeder (the actual data-entry sheet) ──
    ws = wb.active
    ws.title = "KB Feeder"

    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    ws.append(FEEDER_COLUMNS)
    for col_idx in range(1, len(FEEDER_COLUMNS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.append(EXAMPLE_ROW)
    example_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    for col_idx in range(1, len(FEEDER_COLUMNS) + 1):
        ws.cell(row=2, column=col_idx).fill = example_fill

    # Column widths
    widths = [22, 45, 40, 16, 40, 40, 22]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30

    # Dropdown validation for "Precisa usar a Rate Card Lookup Query?"
    tariff_col_idx = FEEDER_COLUMNS.index(COL_NEEDS_TARIFF) + 1
    dv_tariff = DataValidation(type="list", formula1='"Sim,Não"', allow_blank=True)
    ws.add_data_validation(dv_tariff)
    dv_tariff.add(f"{get_column_letter(tariff_col_idx)}2:{get_column_letter(tariff_col_idx)}500")

    # Dropdown validation for Categoria, populated from existing categories if available
    try:
        existing_categories = sorted(
            c for c in get_error_types_df().get("Categoria", []).unique().tolist() if str(c).strip()
        )
    except Exception:
        existing_categories = []
    if existing_categories:
        cat_col_idx = FEEDER_COLUMNS.index(COL_CATEGORY) + 1
        formula = '"' + ",".join(existing_categories)[:255] + '"'
        dv_cat = DataValidation(type="list", formula1=formula, allow_blank=True)
        ws.add_data_validation(dv_cat)
        dv_cat.add(f"{get_column_letter(cat_col_idx)}2:{get_column_letter(cat_col_idx)}500")

    # Add ~200 pre-formatted (but empty) rows so the file "looks" ready to fill
    for r in range(3, 203):
        for c in range(1, len(FEEDER_COLUMNS) + 1):
            ws.cell(row=r, column=c, value="")

    # ── Sheet 2: Instructions ──
    ws2 = wb.create_sheet("Instructions")
    ws2.append(["Knowledge Feeder File — How to fill this in"])
    ws2["A1"].font = Font(bold=True, size=14)
    ws2.append([])
    for row in INSTRUCTIONS:
        ws2.append(list(row))
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 90
    for row_cells in ws2.iter_rows(min_row=3, max_row=3):
        for cell in row_cells:
            cell.font = Font(bold=True)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
