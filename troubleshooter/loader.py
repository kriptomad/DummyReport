import os
import re
import threading
import pandas as pd

ASSETS_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "assets")
STEPS_PATH = os.path.join(ASSETS_DIR, "stepsdummy.xlsx")
FALLBACK_XLSX = os.path.join(ASSETS_DIR, "dummytroubleshoot.xlsx")
FALLBACK_CSV  = os.path.join(ASSETS_DIR, "dummytroubleshoot.csv")

# ── Column names (Main sheet) ────────────────────────────────
COL_CATEGORY      = "Categoria"
COL_ERROR_PATTERN = "Mensagem de erro / padrão identificado"
COL_MEANING       = "Significado provável"
COL_NEEDS_TARIFF  = "Precisa usar a Rate Card Lookup Query?"
COL_HOW_TO_CHECK  = "Como validar"
COL_ACTION        = "Ação recomendada"
COL_RESPONSIBLE   = "Responsável sugerido"

# Optional English-translation columns. Historically these had to be
# filled in manually (no auto-translate). Now, if an AI provider is
# configured (OPENAI_API_KEY / ANTHROPIC_API_KEY / GEMINI_API_KEY —
# same providers used elsewhere in the app), troubleshooter/feedback_store.py
# auto-translates the missing side (PT -> EN or EN -> PT) whenever a KB
# entry is created or edited — see _auto_translate_kb_fields(). If no
# provider is configured, or translation fails, the field is simply left
# blank and the UI falls back to showing the original-language text with
# a note (see app.py's _kb_text_for_lang()).
COL_MEANING_EN       = "Significado provável (English)"
COL_HOW_TO_CHECK_EN  = "Como validar (English)"
COL_ACTION_EN        = "Ação recomendada (English)"

REQUIRED_COLUMNS = [
    COL_ERROR_PATTERN, COL_MEANING, COL_NEEDS_TARIFF,
    COL_HOW_TO_CHECK, COL_ACTION, COL_RESPONSIBLE,
]

# Auto-added if missing, but never required (so old KB files keep working).
OPTIONAL_COLUMNS = [COL_MEANING_EN, COL_HOW_TO_CHECK_EN, COL_ACTION_EN]

TRUTHY_VALUES = {"sim", "yes", "s", "y", "true", "1"}
FALSY_VALUES  = {"nao", "não", "no", "n", "false", "0"}

# ── Column Aliases for Smart Mapping ─────────────────────────
COLUMN_ALIASES = {
    "Categoria": [
        "Category", "Type", "Error Type", "Tipo", "Categoría"
    ],
    "Mensagem de erro / padrão identificado": [
        "Error Message", "Error Pattern", "Error", "Message",
        "Erro", "Mensagem", "Pattern", "Padrão", "Mensagem de erro",
        "Error Msg", "Err Msg", "Error Text"
    ],
    "Significado provável": [
        "Meaning", "Description", "Significado", "Descrição",
        "Probable Meaning", "Desc"
    ],
    "Como validar": [
        "Validation", "How to Validate", "Validação", "How to Check",
        "Validate", "Check"
    ],
    "Ação recomendada": [
        "Action", "Recommended Action", "Solution", "Ação", "Solução",
        "Fix", "Resolution", "Recommended Solution"
    ],
    "Responsável sugerido": [
        "Owner", "Responsible", "Team", "Responsável", "Suggested Owner",
        "Assigned To", "Responsible Team"
    ],
    "Precisa usar a Rate Card Lookup Query?": [
        "Tariff Query", "Needs Tariff Query", "Usa Tariff", "Use Tariff",
        "Rate Card Lookup", "Needs Tariff"
    ],
    "Significado provável (English)": [
        "Meaning (English)", "Meaning EN", "English Meaning", "Meaning_EN",
    ],
    "Como validar (English)": [
        "How to Validate (English)", "How to Check (English)", "Validation EN",
        "Validation (English)", "Check_EN", "How to Validate EN",
    ],
    "Ação recomendada (English)": [
        "Action (English)", "Action EN", "English Action", "Action_EN",
        "Recommended Action (English)", "Solution (English)",
    ],
}

# ── Global caches ────────────────────────────────────────────
_main_df:        pd.DataFrame | None = None
_logic_df:       pd.DataFrame | None = None
_steps_groups:   dict | None = None   # {group_name: [step_text, ...]}
_error_types_df: pd.DataFrame | None = None
_cache_lock = threading.RLock()

def _demo_kb_rows() -> list[dict[str, str]]:
    return [
        {
            COL_CATEGORY: "Connectivity",
            COL_ERROR_PATTERN: "Connection timeout while syncing orders",
            COL_MEANING: "The upstream order service did not respond within the expected window.",
            COL_NEEDS_TARIFF: "No",
            COL_HOW_TO_CHECK: "Check the demo service health panel and confirm the sync queue is backing up.",
            COL_ACTION: "Retry the sync after the service recovers and reprocess the affected order batch.",
            COL_RESPONSIBLE: "Integration Support",
            COL_MEANING_EN: "The upstream order service did not respond within the expected window.",
            COL_HOW_TO_CHECK_EN: "Check the demo service health panel and confirm the sync queue is backing up.",
            COL_ACTION_EN: "Retry the sync after the service recovers and reprocess the affected order batch.",
        },
        {
            COL_CATEGORY: "Data Quality",
            COL_ERROR_PATTERN: "Duplicate invoice reference detected",
            COL_MEANING: "The batch import received the same invoice reference more than once.",
            COL_NEEDS_TARIFF: "No",
            COL_HOW_TO_CHECK: "Review the imported rows and compare duplicate reference values.",
            COL_ACTION: "Remove the duplicate source row and rerun the import job.",
            COL_RESPONSIBLE: "Operations Analyst",
            COL_MEANING_EN: "The batch import received the same invoice reference more than once.",
            COL_HOW_TO_CHECK_EN: "Review the imported rows and compare duplicate reference values.",
            COL_ACTION_EN: "Remove the duplicate source row and rerun the import job.",
        },
        {
            COL_CATEGORY: "Access",
            COL_ERROR_PATTERN: "Authentication token expired for warehouse feed",
            COL_MEANING: "The cached API token used by the warehouse connector is no longer valid.",
            COL_NEEDS_TARIFF: "No",
            COL_HOW_TO_CHECK: "Confirm the token age and validate that refresh calls are failing.",
            COL_ACTION: "Generate a new demo token and restart the warehouse feed connector.",
            COL_RESPONSIBLE: "Platform Team",
            COL_MEANING_EN: "The cached API token used by the warehouse connector is no longer valid.",
            COL_HOW_TO_CHECK_EN: "Confirm the token age and validate that refresh calls are failing.",
            COL_ACTION_EN: "Generate a new demo token and restart the warehouse feed connector.",
        },
    ]


def _ensure_demo_workbook(path: str) -> str:
    if os.path.exists(path):
        return path
    from openpyxl import Workbook

    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb = Workbook()
    ws_main = wb.active
    ws_main.title = "Main"
    ws_main.append([
        COL_CATEGORY, COL_ERROR_PATTERN, COL_MEANING, COL_NEEDS_TARIFF,
        COL_HOW_TO_CHECK, COL_ACTION, COL_RESPONSIBLE,
        COL_MEANING_EN, COL_HOW_TO_CHECK_EN, COL_ACTION_EN,
    ])
    for row in _demo_kb_rows():
        ws_main.append([
            row[COL_CATEGORY], row[COL_ERROR_PATTERN], row[COL_MEANING], row[COL_NEEDS_TARIFF],
            row[COL_HOW_TO_CHECK], row[COL_ACTION], row[COL_RESPONSIBLE],
            row[COL_MEANING_EN], row[COL_HOW_TO_CHECK_EN], row[COL_ACTION_EN],
        ])

    ws_logic = wb.create_sheet("Logic")
    ws_logic.append(["Se a mensagem contém...", "Categoria"])
    ws_logic.append(["timeout", "Connectivity"])
    ws_logic.append(["duplicate", "Data Quality"])
    ws_logic.append(["token", "Access"])

    ws_steps = wb.create_sheet("Steps")
    for line in [
        "Grupo 1 - Validate service availability",
        "Open the demo status dashboard and confirm the upstream service is healthy.",
        "If the service is degraded, wait for recovery before retrying the failed process.",
        "Grupo 2 - Validate source data",
        "Review the imported sample file for duplicates or missing mandatory fields.",
        "Correct the input file and rerun the job in the demo app.",
    ]:
        ws_steps.append([line])

    ws_types = wb.create_sheet("Type of errors")
    ws_types.append(["Category", "Summary"])
    ws_types.append(["Connectivity", "Network or service availability issues."])
    ws_types.append(["Data Quality", "Input or master-data inconsistencies."])
    ws_types.append(["Access", "Credential or permission related failures."])
    wb.save(path)
    return path


# ─────────────────────────────────────────────────────────────
#  INTERNAL HELPERS
# ─────────────────────────────────────────────────────────────

def _safe(v) -> str:
    return "" if v is None else str(v).strip()


def standardize_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    Padroniza nomes de colunas usando aliases.
    Permite carregar arquivos com diferentes formatos de colunas.

    Args:
        df: DataFrame com colunas variadas

    Returns:
        DataFrame com colunas padronizadas
    """
    # Cria mapeamento reverso (alias → nome padrão)
    reverse_map = {}
    for standard_name, aliases in COLUMN_ALIASES.items():
        # O próprio nome padrão também é válido
        reverse_map[standard_name.lower().strip()] = standard_name
        for alias in aliases:
            reverse_map[alias.lower().strip()] = standard_name

    # Identifica colunas a renomear
    renamed_columns = {}
    for col in df.columns:
        col_lower = str(col).lower().strip()
        if col_lower in reverse_map:
            renamed_columns[col] = reverse_map[col_lower]

    # Aplica renomeação
    df = df.rename(columns=renamed_columns)

    # Adiciona colunas faltantes com valores vazios
    for standard_col in REQUIRED_COLUMNS:
        if standard_col not in df.columns:
            df[standard_col] = ""

    # Adiciona categoria se não existir
    if COL_CATEGORY not in df.columns:
        df[COL_CATEGORY] = ""

    # Adiciona colunas opcionais (traduções em inglês) se não existirem
    for optional_col in OPTIONAL_COLUMNS:
        if optional_col not in df.columns:
            df[optional_col] = ""

    return df


def _read_sheet_as_df(path: str, sheet: str) -> pd.DataFrame:
    df = pd.read_excel(path, sheet_name=sheet, header=0, dtype=str, engine="openpyxl")
    df.columns = [_safe(c) for c in df.columns]
    df = df.fillna("")
    return df


def _load_main(path: str) -> pd.DataFrame:
    """Sheet 'Main': error patterns + solutions."""
    df = _read_sheet_as_df(path, "Main")

    # Padroniza colunas para aceitar qualquer formato
    df = standardize_columns(df)

    missing = [c for c in REQUIRED_COLUMNS if c not in df.columns]
    if missing:
        raise ValueError(f"'Main' sheet missing columns: {', '.join(missing)}")

    df = df[df[COL_ERROR_PATTERN].str.strip() != ""].copy()

    tariff_raw = df[COL_NEEDS_TARIFF].str.strip().str.lower()
    df["_needs_tariff"] = tariff_raw.isin(TRUTHY_VALUES)
    df["_pattern_normalized"] = df[COL_ERROR_PATTERN].str.lower().str.strip()

    # Add category from Logic sheet keywords if column absent
    if COL_CATEGORY not in df.columns:
        df[COL_CATEGORY] = ""

    return df.reset_index(drop=True)


def _load_logic(path: str) -> pd.DataFrame:
    """Sheet 'Logic': keyword → category classification."""
    df = _read_sheet_as_df(path, "Logic")
    # Keep rows where keyword column has a value
    kw_col = "Se a mensagem contém..."
    if kw_col in df.columns:
        df = df[df[kw_col].str.strip() != ""].copy()
    return df.reset_index(drop=True)


_STEPS_GROUP_KEYWORDS = ["Grupo", "Se a rate", "Para erros de master data", "Texto resumido", "Operational Error"]
_STEPS_NUMBERED_HEADING_RE = re.compile(r"^\d+\.\s+\S")   # e.g. "2. Se a rate não existir"
_STEPS_STRAY_DIGIT_RE      = re.compile(r"^\d+$")          # leftover merged-cell artifact, e.g. lone "2"
_STEPS_ITEM_NUMBERING_RE   = re.compile(r"^\d+\s+")        # manual "1     text" Excel numbering


def _load_steps(path: str) -> dict:
    """
    Sheet 'Steps': parse into groups {group_name: [steps]}.

    Each cell in the sheet is either:
    - A group/branch title (e.g. "Grupo 1 - Validar rate na query", or a
      numbered branch like "2. Se a rate não existir") — starts a new group.
    - A short sub-heading label (e.g. "Ação:", "Classificar como:") that
      introduces the items right below it within the same group — kept,
      but prefixed with a "### " marker so the UI can render it as a bold
      sub-heading instead of just another numbered step (previously these
      were rendered as flat, confusingly-double-numbered list items).
    - A real step/instruction line, optionally hand-numbered in the sheet
      itself ("1     Rodar a Rate Card Lookup Query.") — that manual numbering
      is stripped here since the UI applies its own numbering.
    - A stray lone digit left over from a merged/wrapped Excel cell — skipped.
    """
    df = pd.read_excel(path, sheet_name="Steps", header=None, dtype=str, engine="openpyxl")
    df = df.fillna("")

    groups: dict[str, list[str]] = {}
    current_group = "Geral"
    current_steps: list[str] = []

    for _, row in df.iterrows():
        cell = _safe(row.iloc[0])
        if not cell:
            continue

        if _STEPS_STRAY_DIGIT_RE.match(cell):
            continue

        is_group_start = (
            any(kw.lower() in cell.lower() for kw in _STEPS_GROUP_KEYWORDS)
            or bool(_STEPS_NUMBERED_HEADING_RE.match(cell))
        )
        if is_group_start:
            if current_steps:
                groups[current_group] = current_steps
            current_group = cell
            current_steps = []
            continue

        is_sub_heading = cell.endswith(":") and len(cell) <= 120 and not cell[0].isdigit()
        if is_sub_heading:
            current_steps.append(f"### {cell[:-1].strip()}")
            continue

        cleaned = _STEPS_ITEM_NUMBERING_RE.sub("", cell).strip()
        if cleaned:
            current_steps.append(cleaned)

    if current_steps:
        groups[current_group] = current_steps

    return groups


def _load_error_types(path: str) -> pd.DataFrame:
    """Sheet 'Type of errors': category metadata."""
    return _read_sheet_as_df(path, "Type of errors")


# ─────────────────────────────────────────────────────────────
#  PUBLIC API
# ─────────────────────────────────────────────────────────────

def load_all(force_reload: bool = False) -> None:
    """Load all sheets from stepsdummy.xlsx into module caches."""
    global _main_df, _logic_df, _steps_groups, _error_types_df

    with _cache_lock:
        if _main_df is not None and not force_reload:
            return

    path = _ensure_demo_workbook(STEPS_PATH)

    new_main_df = _load_main(path)
    new_logic_df = _load_logic(path)
    new_steps_groups = _load_steps(path)
    new_error_types_df = _load_error_types(path)

    with _cache_lock:
        _main_df = new_main_df
        _logic_df = new_logic_df
        _steps_groups = new_steps_groups
        _error_types_df = new_error_types_df


def load_troubleshoot_db(force_reload: bool = False) -> pd.DataFrame:
    """Return the Main matching DataFrame."""
    load_all(force_reload)
    with _cache_lock:
        return _main_df


def get_logic_df() -> pd.DataFrame:
    """Return the Logic classification DataFrame."""
    load_all()
    with _cache_lock:
        return _logic_df


def get_steps_groups() -> dict:
    """Return {group_name: [step_text]} from Steps sheet."""
    load_all()
    with _cache_lock:
        return _steps_groups


def get_error_types_df() -> pd.DataFrame:
    """Return Type of errors DataFrame."""
    load_all()
    with _cache_lock:
        return _error_types_df


# The ERR_MSG column sometimes carries a plain status note even though
# nothing actually went wrong (e.g. "Load created successfully", "Shipment
# processed successfully", "Processo concluído com sucesso", "Completed",
# "Finalizado com sucesso"). These must never be treated as errors —
# matched against the KB, counted in "unique errors" metrics, flagged
# "[!]" in the audit table, or fed into the local AI's learning/gap
# detection — that would be actively misleading and would waste the
# model's vocabulary/training signal on non-errors. `\b` word boundaries
# are relied on to correctly NOT match inside "unsuccessfully"/"insucesso"
# (negated forms) or "incomplete" (a real, different word), since there's
# no boundary between the letters directly abutting the success word in
# those cases.
_SUCCESS_MSG_REGEX = re.compile(
    r"\b("
    r"success(?:ful(?:ly)?)?"          # success, successful, successfully
    r"|sucesso|sucedid[oa]"            # sucesso, sucedido/a
    r"|conclu[íi]d[oa](?:\s+com\s+sucesso)?"  # concluído/a (com sucesso)
    r"|finalizad[oa](?:\s+com\s+sucesso)?"    # finalizado/a (com sucesso)
    r"|complet(?:ed|e|[oa])?"          # complete, completed, completo/a
    r")\b",
    re.IGNORECASE,
)


def is_success_message(err_msg) -> bool:
    """
    Whether `err_msg` describes a successful outcome rather than an actual
    error/failure, and should therefore be excluded from all error
    analysis/counting/highlighting.
    """
    text = str(err_msg or "").strip()
    if not text or text.lower() == "nan":
        return False
    return bool(_SUCCESS_MSG_REGEX.search(text))


def classify_by_keyword(err_msg: str) -> str:
    """Use Logic sheet keywords to classify an ERR_MSG into a category."""
    load_all()
    with _cache_lock:
        logic_df = _logic_df
    kw_col   = "Se a mensagem contém..."
    cat_col  = "Categoria"
    if logic_df is None or kw_col not in logic_df.columns:
        return ""
    err_lower = err_msg.lower()
    for _, row in logic_df.iterrows():
        kw = _safe(row.get(kw_col, "")).lower()
        if kw and kw in err_lower:
            return _safe(row.get(cat_col, ""))
    return ""


def get_next_step_hint(err_msg: str) -> str:
    """
    Use the Logic sheet's 'Próximo passo' column to find a short,
    keyword-specific "what to do next" hint for this exact error message.

    This is more granular than the category-level Steps groups below
    (several distinct keywords can share one category, e.g. "Operational
    Error" covers 3 different messages, each with its own next-step text),
    so it's used as the first, most-specific line of the step-by-step list.
    """
    load_all()
    with _cache_lock:
        logic_df = _logic_df
    kw_col   = "Se a mensagem contém..."
    hint_col = "Próximo passo"
    if logic_df is None or kw_col not in logic_df.columns or hint_col not in logic_df.columns:
        return ""
    err_lower = err_msg.lower()
    for _, row in logic_df.iterrows():
        kw = _safe(row.get(kw_col, "")).lower()
        if kw and kw in err_lower:
            return _safe(row.get(hint_col, ""))
    return ""


def get_steps_for_category(
    category: str,
    next_step_hint: str = "",
    fallback_how_to_check: str = "",
    fallback_action: str = "",
) -> list[str]:
    """
    Return the relevant step-by-step list for a given category.

    Combines (in order):
    1. The keyword-specific "next step" hint (most specific, from Logic sheet).
    2. The detailed multi-step group from the Steps sheet, when the category
       maps to one of the two hand-written groups (rate/lane or master data).
    3. If neither of the above produced real guidance (e.g. categories added
       later — such as "Operational Error" — that were never given a Steps
       sheet group), fall back to the matched KB row's own "how to validate"
       and "recommended action" text, so every category always shows *some*
       tailored step-by-step guidance instead of an empty section.
    """
    load_all()
    steps: list[str] = []
    with _cache_lock:
        steps_groups = _steps_groups

    if next_step_hint:
        steps.append(f"🎯 {next_step_hint}")

    category_lower = category.lower()

    # Map category to group keyword (covers the two categories that have
    # rich, hand-written multi-step guidance in the Steps sheet).
    group_map = {
        "missing rate": "Grupo 1",
        "itinerary":    "Grupo 1",
        "schedule":     "Grupo 1",
        "master data":  "Para erros de master data",
        "inactive":     "Para erros de master data",
        "equipment":    "Para erros de master data",
        "logistics":    "Para erros de master data",
        "division":     "Para erros de master data",
        "missing dc":   "Para erros de master data",
    }

    target_group_kw = None
    for kw, grp in group_map.items():
        if kw in category_lower:
            target_group_kw = grp
            break

    if target_group_kw and steps_groups:
        for group_name, group_steps in steps_groups.items():
            if target_group_kw.lower() in group_name.lower():
                steps.extend(group_steps)
                break

    # Fallback: category has no dedicated Steps-sheet group (or matched
    # nothing beyond the hint above) — build a minimal but real, tailored
    # checklist from the specific KB row that matched this error.
    if len(steps) <= 1:
        if fallback_how_to_check:
            steps.append(f"✅ Validate: {fallback_how_to_check}")
        if fallback_action:
            steps.append(f"➡️ Action: {fallback_action}")

    return steps
