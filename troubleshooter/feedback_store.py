"""
troubleshooter/feedback_store.py — Self-learning feedback loop.

Lets users mark a shipment/error as "corrected" and record what fixed it.
The correction is:
  1. Logged to data/corrections_history.json (audit trail, never lost).
  2. Merged into assets/stepsdummy.xlsx ("Main" sheet) so the troubleshooting
     engine (troubleshooter/engine.py) picks it up on the next match — either
     updating an existing pattern's recommended action or, if the error text
     doesn't match anything known yet, appending a brand-new row.
  3. Optionally enriched using whichever LLM provider is already configured
     elsewhere in the app (OPENAI_API_KEY / ANTHROPIC_API_KEY / GEMINI_API_KEY)
     to turn a short free-text correction into structured guidance (meaning /
     how-to-validate / action). This is best-effort: if no AI provider is
     configured, a rule-based fallback keeps the feature fully usable offline.
"""
import os
import json
import re
import shutil
from pathlib import Path
from datetime import datetime
from typing import Optional, Dict, Any, List

import pandas as pd
from filelock import FileLock, Timeout

from troubleshooter.loader import (
    STEPS_PATH,
    COL_CATEGORY,
    COL_ERROR_PATTERN,
    COL_MEANING,
    COL_NEEDS_TARIFF,
    COL_HOW_TO_CHECK,
    COL_ACTION,
    COL_RESPONSIBLE,
    COL_MEANING_EN,
    COL_HOW_TO_CHECK_EN,
    COL_ACTION_EN,
    OPTIONAL_COLUMNS,
    classify_by_keyword,
    load_all,
)
from troubleshooter.engine import _match_score
from troubleshooter import kb_ownership
from utils.safe_json import json_transaction
from i18n import t

DATA_DIR = Path(__file__).resolve().parent.parent / "data"
DATA_DIR.mkdir(exist_ok=True)
CORRECTIONS_LOG_PATH = DATA_DIR / "corrections_history.json"

# Serializes every read-modify-write cycle against stepsdummy.xlsx (the KB
# workbook) so two concurrent edits (e.g. two analysts submitting a
# correction at the same time) can't silently clobber each other — pandas/
# openpyxl have no locking of their own. Also guards the backup-before-write
# below.
_STEPS_LOCK_PATH = str(STEPS_PATH) + ".lock"
_STEPS_LOCK_TIMEOUT_SECONDS = 15
_KB_BACKUPS_DIR = Path(__file__).resolve().parent.parent / "assets" / "knowledge_base_backups"
_KB_BACKUPS_KEEP = 20

STRONG_MATCH_THRESHOLD = 0.75


def _backup_steps_file() -> None:
    """Copies the current stepsdummy.xlsx to a timestamped backup before it
    gets overwritten, so a bad/corrupt merge is always recoverable. Best
    effort — never blocks the actual save on backup failure."""
    try:
        if not os.path.exists(STEPS_PATH):
            return
        _KB_BACKUPS_DIR.mkdir(parents=True, exist_ok=True)
        stamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        backup_path = _KB_BACKUPS_DIR / f"kb_feedback_{stamp}.xlsx"
        shutil.copy2(STEPS_PATH, backup_path)
        backups = sorted(_KB_BACKUPS_DIR.glob("kb_feedback_*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
        for old in backups[_KB_BACKUPS_KEEP:]:
            try:
                old.unlink()
            except OSError:
                pass
    except Exception:
        pass


def _lock_steps_file() -> FileLock:
    return FileLock(_STEPS_LOCK_PATH, timeout=_STEPS_LOCK_TIMEOUT_SECONDS)


# ── Correction audit log (JSON, never lost even if KB write fails) ──────────

def _load_corrections_log() -> list:
    if not CORRECTIONS_LOG_PATH.exists():
        return []
    try:
        with open(CORRECTIONS_LOG_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return []


def get_corrections_history(limit: int = 50) -> List[Dict[str, Any]]:
    """Return the most recent correction/feedback entries (newest first)."""
    entries = _load_corrections_log()
    return list(reversed(entries))[:limit]


# ── AI enrichment (best-effort, provider-agnostic) ───────────────────────────

def _parse_json_loose(raw: str) -> Optional[dict]:
    raw = (raw or "").strip().replace("```json", "").replace("```", "").strip()
    try:
        return json.loads(raw)
    except Exception:
        match = re.search(r"\{.*\}", raw, re.DOTALL)
        if match:
            try:
                return json.loads(match.group(0))
            except Exception:
                return None
    return None


def _ai_call_openai(prompt: str) -> str:
    import openai
    client = openai.OpenAI(api_key=os.getenv("OPENAI_API_KEY"))
    response = client.chat.completions.create(
        model="gpt-4o",
        messages=[
            {"role": "system", "content": "You produce concise, structured troubleshooting knowledge-base entries. Reply with JSON only."},
            {"role": "user", "content": prompt},
        ],
        temperature=0.2,
        max_tokens=400,
    )
    return response.choices[0].message.content.strip()


def _ai_call_anthropic(prompt: str) -> str:
    import anthropic
    client = anthropic.Anthropic(api_key=os.getenv("ANTHROPIC_API_KEY"))
    response = client.messages.create(
        model="claude-3-5-sonnet-latest",
        max_tokens=400,
        messages=[{"role": "user", "content": prompt}],
    )
    return response.content[0].text.strip()


def _ai_call_gemini(prompt: str) -> str:
    from google import genai
    from google.genai import types
    client = genai.Client(api_key=os.getenv("GEMINI_API_KEY"))
    response = client.models.generate_content(
        model="gemini-2.5-flash",
        contents=prompt,
        config=types.GenerateContentConfig(temperature=0.2, max_output_tokens=400),
    )
    return (response.text or "").strip()


def _enrich_with_ai(err_msg: str, correction_text: str) -> Dict[str, str]:
    """
    Best-effort: ask whichever LLM provider is configured to turn a short
    free-text correction into structured KB fields. Falls back to simple
    heuristics if no provider is available or every call fails — the
    feature always works, AI just makes the entry nicer.
    """
    prompt = f"""You are helping build a troubleshooting knowledge base for shipment audit errors.

Error message: "{err_msg}"
Correction applied by a human analyst: "{correction_text}"

Based on this, produce a JSON object with exactly these keys:
- "meaning": one sentence — the likely root cause of this error.
- "how_to_validate": one or two short sentences on how to confirm this is the issue.
- "action": one or two short sentences — the recommended fix (base this on the correction given).

Respond with ONLY the JSON object, no markdown, no extra text.
"""
    providers = []
    if os.getenv("OPENAI_API_KEY"):
        providers.append(("openai", _ai_call_openai))
    if os.getenv("ANTHROPIC_API_KEY"):
        providers.append(("anthropic", _ai_call_anthropic))
    if os.getenv("GEMINI_API_KEY"):
        providers.append(("gemini", _ai_call_gemini))

    for name, fn in providers:
        try:
            raw = fn(prompt)
            data = _parse_json_loose(raw)
            if data:
                return {
                    "meaning": (data.get("meaning") or "").strip() or f"Learned from user correction: {correction_text[:120]}",
                    "how_to_validate": (data.get("how_to_validate") or "").strip() or "Review the shipment details reported by the user.",
                    "action": (data.get("action") or "").strip() or correction_text,
                    "ai_provider": name,
                }
        except Exception:
            continue

    # Rule-based fallback (no AI configured / all providers failed)
    return {
        "meaning": "Pattern learned from a user-reported correction.",
        "how_to_validate": "Compare against the original error message and the correction notes below.",
        "action": correction_text,
        "ai_provider": "none",
    }


# ── AI translation (best-effort, provider-agnostic) ──────────────────────────
# Reuses the same LLM providers as _enrich_with_ai() above. Used to keep the
# bilingual KB columns (Significado provável / (English), Como validar /
# (English), Ação recomendada / (English)) in sync automatically: whichever
# language a KB entry is written in, the other language gets auto-filled on
# save. Best-effort — if no provider is configured, or every call fails,
# the translated field is simply left blank (existing, pre-AI behavior).

_LANG_NAMES = {"pt": "Brazilian Portuguese", "en": "English"}


def translate_text(text: str, target_lang: str) -> str:
    """Public wrapper around _ai_translate(), for callers outside this
    module (e.g. app.py's on-the-fly display translation) that need a
    single piece of text translated without going through a full KB row."""
    return _ai_translate(text, target_lang)


try:
    from deep_translator import GoogleTranslator as _GoogleTranslator
except ImportError:  # pragma: no cover
    _GoogleTranslator = None


def _ai_translate(text: str, target_lang: str) -> str:
    """Translates `text` into `target_lang` ('pt' or 'en'). Tries whichever
    LLM provider is configured first (OpenAI/Anthropic/Gemini), then falls
    back to a free, no-API-key translation service (Google Translate via
    `deep-translator`) so this always works even when no paid AI provider
    key is set — which, as of writing, is the actual state of this
    deployment (no OPENAI_API_KEY/ANTHROPIC_API_KEY/GEMINI_API_KEY
    configured). Returns "" only if every option fails/is unavailable
    (e.g. no internet access) — callers must treat that as "no
    translation, leave field blank"."""
    text = (text or "").strip()
    if not text:
        return ""

    target_name = _LANG_NAMES.get(target_lang, "English")
    prompt = (
        f"Translate the following shipment-troubleshooting knowledge-base text "
        f"into {target_name}. Keep it concise and preserve any technical terms, "
        f"codes, or field names as-is. Reply with ONLY the translated text, no "
        f"quotes, no extra commentary.\n\nText:\n{text}"
    )

    providers = []
    if os.getenv("OPENAI_API_KEY"):
        providers.append(_ai_call_openai)
    if os.getenv("ANTHROPIC_API_KEY"):
        providers.append(_ai_call_anthropic)
    if os.getenv("GEMINI_API_KEY"):
        providers.append(_ai_call_gemini)

    for fn in providers:
        try:
            translated = fn(prompt).strip()
            if translated:
                return translated
        except Exception:
            continue

    # Free fallback — no API key required. Used whenever no LLM provider
    # is configured (the common case here) or every configured provider
    # call failed. Best-effort: any network/library error just means "no
    # translation available", never raises up to the caller.
    if _GoogleTranslator is not None:
        try:
            source = "en" if target_lang == "pt" else "pt"
            translated = _GoogleTranslator(source=source, target=target_lang).translate(text)
            if translated and translated.strip():
                return translated.strip()
        except Exception:
            pass
    return ""


def _ensure_optional_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Adds the bilingual English-translation columns to `df` if this
    workbook predates them, so assigning to them below never silently
    no-ops just because the column didn't exist yet."""
    for col in OPTIONAL_COLUMNS:
        if col not in df.columns:
            df[col] = ""
    return df


def _auto_translate_kb_row(df_main: pd.DataFrame, idx, cws: Optional[str] = None) -> None:
    """
    Best-effort bilingual sync for a single KB row (in place): for each of
    the 3 translatable field pairs, if one language's text is present and
    the other is blank, auto-translate and fill the blank side. Silently
    does nothing if no AI provider is configured — the KB remains fully
    usable with manually-entered or blank translations, same as before.
    """
    pairs = [
        (COL_MEANING, COL_MEANING_EN),
        (COL_HOW_TO_CHECK, COL_HOW_TO_CHECK_EN),
        (COL_ACTION, COL_ACTION_EN),
    ]
    for pt_col, en_col in pairs:
        if pt_col not in df_main.columns or en_col not in df_main.columns:
            continue
        pt_text = str(df_main.at[idx, pt_col] or "").strip()
        en_text = str(df_main.at[idx, en_col] or "").strip()
        if pt_text and not en_text:
            translated = _ai_translate(pt_text, "en")
            if translated:
                df_main.at[idx, en_col] = translated
        elif en_text and not pt_text:
            translated = _ai_translate(en_text, "pt")
            if translated:
                df_main.at[idx, pt_col] = translated


def backfill_kb_translations() -> Dict[str, Any]:
    """
    One-shot batch job: goes through every row in the KB and fills in any
    blank EN/PT translation using `_auto_translate_kb_row` — fixes legacy
    KB entries that were created before the bilingual auto-translate
    feature existed (their EN columns are blank, so until now the app
    fell back to showing Portuguese text even when the UI language was
    English). Safe to re-run any time; already-translated rows are
    left untouched (only truly blank sides get filled).

    Returns {"ok": True, "rows_updated": int, "rows_total": int} or
    {"ok": False, "reason": str} if the KB file couldn't be locked/read.
    """
    if not os.path.exists(STEPS_PATH):
        return {"ok": False, "reason": t("feedback.kb_not_found", path=STEPS_PATH)}

    try:
        with _lock_steps_file():
            df_main = pd.read_excel(STEPS_PATH, sheet_name="Main", header=0, dtype=str, engine="openpyxl").fillna("")
            df_main = _ensure_optional_columns(df_main)

            pairs = [
                (COL_MEANING, COL_MEANING_EN),
                (COL_HOW_TO_CHECK, COL_HOW_TO_CHECK_EN),
                (COL_ACTION, COL_ACTION_EN),
            ]
            rows_updated = 0
            for idx in df_main.index:
                before = tuple(str(df_main.at[idx, en_col] or "") for _, en_col in pairs if en_col in df_main.columns)
                _auto_translate_kb_row(df_main, idx)
                after = tuple(str(df_main.at[idx, en_col] or "") for _, en_col in pairs if en_col in df_main.columns)
                if before != after:
                    rows_updated += 1

            if rows_updated:
                _write_main_sheet(df_main)

            return {"ok": True, "rows_updated": rows_updated, "rows_total": len(df_main)}
    except Timeout:
        return {"ok": False, "reason": t("feedback.kb_busy")}


# ── Knowledge-base merge (assets/stepsdummy.xlsx, "Main" sheet) ─────────────

def _write_main_sheet(df_main: pd.DataFrame) -> None:
    """Writes the updated Main sheet back to stepsdummy.xlsx, preserving all
    other sheets. Backs up the previous version first, and writes via a
    temp-file + atomic rename so a crash/interruption mid-save can never
    leave a half-written/corrupt workbook behind. Caller MUST already hold
    `_lock_steps_file()` for the whole read-modify-write cycle."""
    from openpyxl import load_workbook
    from openpyxl.utils.dataframe import dataframe_to_rows

    _backup_steps_file()

    wb = load_workbook(STEPS_PATH)
    if "Main" in wb.sheetnames:
        idx = wb.sheetnames.index("Main")
        del wb["Main"]
    else:
        idx = 0
    ws = wb.create_sheet("Main", idx)

    for row in dataframe_to_rows(df_main, index=False, header=True):
        ws.append(row)

    tmp_path = str(STEPS_PATH) + ".tmp"
    try:
        wb.save(tmp_path)
        os.replace(tmp_path, STEPS_PATH)  # atomic on Windows & POSIX
    finally:
        if os.path.exists(tmp_path):
            try:
                os.remove(tmp_path)
            except OSError:
                pass


def delete_kb_entry(pattern: str, cws: str, force: bool = False) -> Dict[str, Any]:
    """
    Deletes a knowledge-base row (exact match on COL_ERROR_PATTERN) from the
    'Main' sheet, provided `cws` is the entry's owner (or `force=True` for
    admin use). Also removes its ownership/history metadata.

    Returns {"deleted": bool, "reason": str (if not deleted)}.
    """
    if not os.path.exists(STEPS_PATH):
        raise FileNotFoundError(t("feedback.kb_not_found", path=STEPS_PATH))

    if not force and not kb_ownership.can_edit_directly(pattern, cws):
        owner = kb_ownership.get_owner(pattern)
        return {"deleted": False, "reason": t("feedback.only_owner_delete", owner=owner)}

    # Hold the KB file lock across the whole read → mutate → write cycle so
    # two concurrent edits can't interleave and silently lose one of them.
    try:
        with _lock_steps_file():
            df_main = pd.read_excel(STEPS_PATH, sheet_name="Main", header=0, dtype=str, engine="openpyxl").fillna("")
            if COL_ERROR_PATTERN not in df_main.columns:
                return {"deleted": False, "reason": t("feedback.no_pattern_column")}

            mask = df_main[COL_ERROR_PATTERN].str.strip().str.lower() == pattern.strip().lower()
            if not mask.any():
                return {"deleted": False, "reason": t("feedback.entry_not_found")}

            df_main = df_main[~mask].reset_index(drop=True)
            _write_main_sheet(df_main)
    except Timeout:
        return {"deleted": False, "reason": t("feedback.kb_busy")}

    kb_ownership.delete_meta(pattern)
    load_all(force_reload=True)

    return {"deleted": True}


def update_kb_entry(
    pattern: str,
    cws: str,
    new_pattern: Optional[str] = None,
    meaning: Optional[str] = None,
    how_to_check: Optional[str] = None,
    action: Optional[str] = None,
    responsible: Optional[str] = None,
    category: Optional[str] = None,
    needs_tariff: Optional[str] = None,
    meaning_en: Optional[str] = None,
    how_to_check_en: Optional[str] = None,
    action_en: Optional[str] = None,
    force: bool = False,
    auto_translate: bool = True,
) -> Dict[str, Any]:
    """
    Directly edits an existing knowledge-base row (exact match on
    COL_ERROR_PATTERN), provided `cws` is the entry's owner (or SYSTEM-owned,
    or `force=True` for admin use). Any field left as None is left unchanged.

    `new_pattern`, if given, renames the error pattern itself (ownership
    metadata is moved to the new key so history/freshness aren't lost).

    If `auto_translate` is True (default), after applying the given edits
    any translatable field pair (meaning/meaning_en, how_to_check/
    how_to_check_en, action/action_en) that ends up with one language
    filled and the other blank gets the blank side auto-translated via
    whichever AI provider is configured — see _auto_translate_kb_row().

    Returns {"updated": bool, "reason": str (if not updated)}.
    """
    if not os.path.exists(STEPS_PATH):
        raise FileNotFoundError(t("feedback.kb_not_found", path=STEPS_PATH))

    if not force and not kb_ownership.can_edit_directly(pattern, cws):
        owner = kb_ownership.get_owner(pattern)
        return {"updated": False, "reason": t("feedback.only_owner_edit", owner=owner)}

    try:
        with _lock_steps_file():
            df_main = pd.read_excel(STEPS_PATH, sheet_name="Main", header=0, dtype=str, engine="openpyxl").fillna("")
            if COL_ERROR_PATTERN not in df_main.columns:
                return {"updated": False, "reason": t("feedback.no_pattern_column")}

            mask = df_main[COL_ERROR_PATTERN].str.strip().str.lower() == pattern.strip().lower()
            if not mask.any():
                return {"updated": False, "reason": t("feedback.entry_not_found")}
            idx = df_main[mask].index[0]
            df_main = _ensure_optional_columns(df_main)

            existing_action = str(df_main.at[idx, COL_ACTION]) if COL_ACTION in df_main.columns else ""

            final_pattern = (new_pattern or "").strip() or pattern
            df_main.at[idx, COL_ERROR_PATTERN] = final_pattern
            if meaning is not None and COL_MEANING in df_main.columns:
                df_main.at[idx, COL_MEANING] = meaning
            if how_to_check is not None and COL_HOW_TO_CHECK in df_main.columns:
                df_main.at[idx, COL_HOW_TO_CHECK] = how_to_check
            if action is not None and COL_ACTION in df_main.columns:
                df_main.at[idx, COL_ACTION] = action
            if responsible is not None and COL_RESPONSIBLE in df_main.columns:
                df_main.at[idx, COL_RESPONSIBLE] = responsible
            if category is not None and COL_CATEGORY in df_main.columns:
                df_main.at[idx, COL_CATEGORY] = category
            if needs_tariff is not None and COL_NEEDS_TARIFF in df_main.columns:
                df_main.at[idx, COL_NEEDS_TARIFF] = needs_tariff
            if meaning_en is not None and COL_MEANING_EN in df_main.columns:
                df_main.at[idx, COL_MEANING_EN] = meaning_en
            if how_to_check_en is not None and COL_HOW_TO_CHECK_EN in df_main.columns:
                df_main.at[idx, COL_HOW_TO_CHECK_EN] = how_to_check_en
            if action_en is not None and COL_ACTION_EN in df_main.columns:
                df_main.at[idx, COL_ACTION_EN] = action_en

            if auto_translate:
                _auto_translate_kb_row(df_main, idx, cws)

            _write_main_sheet(df_main)
    except Timeout:
        return {"updated": False, "reason": t("feedback.kb_busy")}

    if final_pattern.strip().lower() != pattern.strip().lower():
        kb_ownership.rename_pattern(pattern, final_pattern, cws or kb_ownership.SYSTEM_OWNER)
    else:
        kb_ownership.stamp_updated(final_pattern, cws or kb_ownership.SYSTEM_OWNER, previous_action_snapshot=existing_action)

    load_all(force_reload=True)
    return {"updated": True, "pattern": final_pattern}


def find_best_match_pattern(err_msg: str) -> Dict[str, Any]:
    """
    Finds the best-matching Main-sheet row for err_msg without mutating
    anything. Used by callers (e.g. the UI) that need to know who OWNS the
    entry that would be affected before deciding whether to apply a direct
    edit or route through the fix-request workflow.

    Returns {"pattern": str|None, "score": float, "row": int|None}.
    """
    if not os.path.exists(STEPS_PATH):
        return {"pattern": None, "score": 0.0, "row": None}

    df_main = pd.read_excel(STEPS_PATH, sheet_name="Main", header=0, dtype=str, engine="openpyxl").fillna("")
    err_normalized = err_msg.lower().strip()

    best_idx = None
    best_score = 0.0
    best_pattern = None
    if COL_ERROR_PATTERN in df_main.columns:
        for idx, row in df_main.iterrows():
            pattern = str(row.get(COL_ERROR_PATTERN, "")).strip()
            if not pattern:
                continue
            score = _match_score(err_normalized, pattern.lower())
            if score > best_score:
                best_score = score
                best_idx = idx
                best_pattern = pattern

    return {"pattern": best_pattern, "score": round(best_score, 3), "row": int(best_idx) if best_idx is not None else None}


def _merge_into_knowledge_base(
    err_msg: str,
    meaning: str,
    how_to_validate: str,
    action_text: str,
    category: str,
    responsible: str,
    cws: Optional[str] = None,
) -> Dict[str, Any]:
    """
    Finds the best-matching row in 'Main' for err_msg. If the match is
    strong, updates its 'Ação recomendada' / 'Como validar' /
    'Responsável sugerido' fields (appending, so history isn't lost). If no
    strong match exists, appends a brand-new row — this is how the app
    "learns" new troubleshooting entries from real-world corrections.

    `cws` is the CWS of the user submitting the change; it is stamped as the
    entry's owner via troubleshooter.kb_ownership (defaults to "SYSTEM" if
    not provided, e.g. for automated/legacy calls).
    """
    if not os.path.exists(STEPS_PATH):
        raise FileNotFoundError(t("feedback.kb_not_found", path=STEPS_PATH))

    try:
        with _lock_steps_file():
            df_main = pd.read_excel(STEPS_PATH, sheet_name="Main", header=0, dtype=str, engine="openpyxl").fillna("")
            df_main = _ensure_optional_columns(df_main)

            err_normalized = err_msg.lower().strip()

            best_idx = None
            best_score = 0.0
            if COL_ERROR_PATTERN in df_main.columns:
                for idx, row in df_main.iterrows():
                    pattern = str(row.get(COL_ERROR_PATTERN, "")).lower().strip()
                    if not pattern:
                        continue
                    score = _match_score(err_normalized, pattern)
                    if score > best_score:
                        best_score = score
                        best_idx = idx

            if best_idx is not None and best_score >= STRONG_MATCH_THRESHOLD:
                existing_pattern = str(df_main.at[best_idx, COL_ERROR_PATTERN])
                existing_action = str(df_main.at[best_idx, COL_ACTION]) if COL_ACTION in df_main.columns else ""
                stamp = datetime.now().strftime("%Y-%m-%d")
                appended = f"{existing_action}\n\n[Updated {stamp}] {action_text}".strip()
                df_main.at[best_idx, COL_ACTION] = appended
                if COL_HOW_TO_CHECK in df_main.columns and not str(df_main.at[best_idx, COL_HOW_TO_CHECK]).strip():
                    df_main.at[best_idx, COL_HOW_TO_CHECK] = how_to_validate
                if COL_RESPONSIBLE in df_main.columns and not str(df_main.at[best_idx, COL_RESPONSIBLE]).strip():
                    df_main.at[best_idx, COL_RESPONSIBLE] = responsible
                action = "updated"
                row_ref = best_idx
                matched_pattern = existing_pattern
                kb_ownership.stamp_updated(existing_pattern, cws or kb_ownership.SYSTEM_OWNER, previous_action_snapshot=existing_action)
                # The action text changed (appended) — refresh its English
                # translation too, but leave meaning/how-to-check EN alone
                # since those weren't touched on an "updated" match.
                if COL_ACTION_EN in df_main.columns:
                    translated = _ai_translate(appended, "en")
                    if translated:
                        df_main.at[best_idx, COL_ACTION_EN] = translated
            else:
                new_row = {c: "" for c in df_main.columns}
                new_row[COL_ERROR_PATTERN] = err_msg
                new_row[COL_MEANING] = meaning
                new_row[COL_NEEDS_TARIFF] = "Não"
                new_row[COL_HOW_TO_CHECK] = how_to_validate
                new_row[COL_ACTION] = action_text
                new_row[COL_RESPONSIBLE] = responsible
                if COL_CATEGORY in df_main.columns:
                    new_row[COL_CATEGORY] = category
                # Brand-new KB entry: auto-translate PT -> EN right away so
                # the new fix is immediately readable in either language —
                # this is the "adding new knowledge auto-translates" feature.
                new_row[COL_MEANING_EN] = _ai_translate(meaning, "en")
                new_row[COL_HOW_TO_CHECK_EN] = _ai_translate(how_to_validate, "en")
                new_row[COL_ACTION_EN] = _ai_translate(action_text, "en")
                df_main = pd.concat([df_main, pd.DataFrame([new_row])], ignore_index=True)
                action = "created"
                row_ref = len(df_main) - 1
                matched_pattern = None
                kb_ownership.stamp_created(err_msg, cws or kb_ownership.SYSTEM_OWNER)

            _write_main_sheet(df_main)
    except Timeout:
        return {"action": "failed", "row": None, "match_score": 0.0, "reason": t("feedback.kb_busy")}

    return {"action": action, "row": int(row_ref), "match_score": round(best_score, 3), "matched_pattern": matched_pattern}


def suggest_new_kb_fix(err_msg: str) -> Dict[str, Any]:
    """
    Best-effort AI-assisted draft fix for a brand-new, previously-unseen
    error message that had no KB match — used by the "Pendências" tab so
    an analyst gets a starting point instead of a blank form. Reuses the
    Troubleshooter's own matching (troubleshooter.engine.match_errors) to
    find the closest existing KB entry as context, then asks whichever
    LLM provider is configured (same priority order as _enrich_with_ai)
    to draft meaning/how-to-validate/action/category for the new error.

    IMPORTANT: this NEVER writes to the KB — it only returns a suggestion
    for a human to review, edit and explicitly approve (see
    troubleshooter.pending_errors.approve()).

    Returns {"suggestion": {...}, "similar": {...}|None, "source": str}
    where `source` is the AI provider name, "similar_kb" (no AI
    configured/available, fell back to the closest KB entry's text as a
    draft), or "none" (nothing usable found).
    """
    import pandas as pd
    from troubleshooter.engine import match_errors

    err_msg = str(err_msg or "").strip()
    results = match_errors(pd.Series([err_msg])) if err_msg else []
    similar = None
    if results and results[0]["matches"]:
        top = results[0]["matches"][0]
        # Always surface the similar entry's text in English — this
        # worklist is reviewed by the technical team and shown to
        # business stakeholders, so the draft must never silently be in
        # Portuguese just because that's the KB's canonical language.
        # Falls back to on-the-fly translation if the EN column is blank
        # (legacy KB rows predating the bilingual columns).
        meaning_pt = str(top.get(COL_MEANING, ""))
        how_to_pt = str(top.get(COL_HOW_TO_CHECK, ""))
        action_pt = str(top.get(COL_ACTION, ""))
        meaning_en = str(top.get(COL_MEANING_EN, "")).strip() or _ai_translate(meaning_pt, "en")
        how_to_en = str(top.get(COL_HOW_TO_CHECK_EN, "")).strip() or _ai_translate(how_to_pt, "en")
        action_en = str(top.get(COL_ACTION_EN, "")).strip() or _ai_translate(action_pt, "en")
        similar = {
            "pattern": str(top.get(COL_ERROR_PATTERN, "")),
            "score": float(top.get("_match_score", 0.0)),
            "meaning": meaning_en or meaning_pt,
            "how_to_check": how_to_en or how_to_pt,
            "action": action_en or action_pt,
            "responsible": str(top.get(COL_RESPONSIBLE, "")),
            "category": str(top.get(COL_CATEGORY, "")),
        }

    providers = []
    if os.getenv("OPENAI_API_KEY"):
        providers.append(("openai", _ai_call_openai))
    if os.getenv("ANTHROPIC_API_KEY"):
        providers.append(("anthropic", _ai_call_anthropic))
    if os.getenv("GEMINI_API_KEY"):
        providers.append(("gemini", _ai_call_gemini))

    context = ""
    if similar and similar["score"] >= 0.3:
        context = (
            f'A somewhat similar known pattern already exists in the knowledge base '
            f'(similarity {similar["score"]:.0%}): pattern "{similar["pattern"]}", '
            f'meaning "{similar["meaning"]}", action "{similar["action"]}". '
            f'Use it only as context — tailor your answer to the NEW error below, '
            f"don't just copy it verbatim.\n\n"
        )
    prompt = (
        "You are helping build a troubleshooting knowledge base for shipment audit errors.\n\n"
        f"{context}"
        f'New, previously unseen error message: "{err_msg}"\n\n'
        "Produce a JSON object with exactly these keys:\n"
        '- "meaning": one sentence — the likely root cause of this error.\n'
        '- "how_to_validate": one or two short sentences on how to confirm this is the issue.\n'
        '- "action": one or two short sentences — the recommended fix.\n'
        '- "category": a short category label (2-4 words).\n\n'
        "Respond with ONLY the JSON object, no markdown, no extra text.\n"
    )

    for name, fn in providers:
        try:
            raw = fn(prompt)
            data = _parse_json_loose(raw)
            if data and ((data.get("meaning") or "").strip() or (data.get("action") or "").strip()):
                suggestion = {
                    "meaning": (data.get("meaning") or "").strip(),
                    "how_to_check": (data.get("how_to_validate") or "").strip(),
                    "action": (data.get("action") or "").strip(),
                    "category": (data.get("category") or (similar or {}).get("category", "")).strip(),
                    "responsible": (similar or {}).get("responsible", ""),
                }
                return {"suggestion": suggestion, "similar": similar, "source": name}
        except Exception:
            continue

    if similar and similar["score"] >= 0.5:
        # No AI provider configured (or every call failed) but a decently
        # similar KB entry exists — offer its text as an editable
        # starting point instead of leaving the analyst with a blank form.
        suggestion = {
            "meaning": similar["meaning"],
            "how_to_check": similar["how_to_check"],
            "action": similar["action"],
            "responsible": similar["responsible"],
            "category": similar["category"],
        }
        return {"suggestion": suggestion, "similar": similar, "source": "similar_kb"}

    return {
        "suggestion": {"meaning": "", "how_to_check": "", "action": "", "responsible": "", "category": ""},
        "similar": similar,
        "source": "none",
    }


def create_kb_entry_from_pending(
    err_msg: str,
    meaning: str,
    how_to_check: str,
    action: str,
    cws: str,
    responsible: str = "",
    category: str = "",
) -> Dict[str, Any]:
    """
    Public wrapper around _merge_into_knowledge_base() for the "Pendências"
    approval flow (troubleshooter/pending_errors.py): posts a
    human-reviewed (and possibly AI-drafted/edited) fix to the KB for a
    previously-unmatched error, stamping `cws` as the entry's owner.
    """
    return _merge_into_knowledge_base(
        err_msg=err_msg,
        meaning=meaning,
        how_to_validate=how_to_check,
        action_text=action,
        category=category,
        responsible=responsible,
        cws=cws,
    )


# ── Public API ───────────────────────────────────────────────────────────────

def submit_correction(
    shipment_id: str,
    err_msg: str,
    correction_text: str,
    corrected: bool = True,
    category: Optional[str] = None,
    responsible: Optional[str] = None,
    use_ai: bool = True,
    cws: Optional[str] = None,
    user_name: Optional[str] = None,
    force_direct: bool = False,
) -> Dict[str, Any]:
    """
    Records a user-provided correction for a given shipment/error and teaches
    the troubleshooting knowledge base (assets/stepsdummy.xlsx, 'Main' sheet).

    Ownership gating: if the best-matching KB row is already owned by a
    DIFFERENT user (not SYSTEM, not `cws`), the change is NOT applied
    directly — instead a pending fix-request is created for the row's owner
    to review (see troubleshooter.fix_requests). Pass `force_direct=True` to
    bypass this (e.g. for admin/system use).

    Returns a dict describing what happened, e.g.
    {"action": "created"|"updated"|"logged_only"|"pending_request", "row": int, "ai_provider": str}
    """
    err_msg = str(err_msg or "").strip()
    correction_text = str(correction_text or "").strip()

    log_entry = {
        "shipment_id": shipment_id,
        "err_msg": err_msg,
        "corrected": corrected,
        "correction_text": correction_text,
        "category": category,
        "timestamp": datetime.now().isoformat(timespec="seconds"),
    }

    if not corrected or not correction_text:
        # Just log the feedback (e.g. "not corrected yet"), no KB mutation.
        with json_transaction(CORRECTIONS_LOG_PATH, default=[]) as entries:
            entries.append(log_entry)
            del entries[:-1000]
        return {"action": "logged_only"}

    ai_fields = _enrich_with_ai(err_msg, correction_text) if use_ai else None

    meaning = (ai_fields or {}).get("meaning") or f"Learned from correction: {correction_text[:150]}"
    how_to_validate = (ai_fields or {}).get("how_to_validate") or "Review shipment details reported by the analyst."
    action_text = (ai_fields or {}).get("action") or correction_text
    resolved_category = category or classify_by_keyword(err_msg) or "Learned / User-reported"

    # ── Ownership gate: is there an existing, strongly-matching row owned by
    #    someone else? If so, route through the request/approval workflow
    #    instead of silently overwriting another user's fix.
    if not force_direct and cws:
        match = find_best_match_pattern(err_msg)
        if match["pattern"] and match["score"] >= STRONG_MATCH_THRESHOLD:
            if not kb_ownership.can_edit_directly(match["pattern"], cws):
                from troubleshooter.fix_requests import create_request
                owner_cws = kb_ownership.get_owner(match["pattern"])
                req = create_request(
                    requester_cws=cws,
                    requester_name=user_name or cws,
                    owner_cws=owner_cws,
                    err_pattern=match["pattern"],
                    request_type="improvement",
                    message=f"Suggested correction for shipment {shipment_id}: {correction_text}",
                    proposed_action=action_text,
                )
                log_entry["kb_action"] = "pending_request"
                with json_transaction(CORRECTIONS_LOG_PATH, default=[]) as entries:
                    entries.append(log_entry)
                    del entries[:-1000]
                return {"action": "pending_request", "request_id": req["id"], "owner_cws": owner_cws}

    result = _merge_into_knowledge_base(
        err_msg=err_msg,
        meaning=meaning,
        how_to_validate=how_to_validate,
        action_text=action_text,
        category=resolved_category,
        responsible=responsible or "Learned (auto)",
        cws=cws,
    )

    log_entry["kb_action"] = result["action"]
    log_entry["ai_provider"] = (ai_fields or {}).get("ai_provider", "none")
    log_entry["cws"] = cws
    # Confirmed (err_msg -> KB pattern) label for Level 2's supervised
    # ensemble (troubleshooter/ilt_ai_core.py) — only meaningful when this
    # correction was actually attached to an EXISTING KB entry (an
    # implicit "yes, this is the right match" confirmation), not when a
    # brand-new row was created (there's no "match" to learn from yet in
    # that case, just like PSLD's queue only trains on confirmed matches,
    # never on brand-new KB creations).
    if result["action"] == "updated" and result.get("matched_pattern"):
        log_entry["matched_pattern"] = result["matched_pattern"]
    with json_transaction(CORRECTIONS_LOG_PATH, default=[]) as entries:
        entries.append(log_entry)
        del entries[:-1000]

    # Force the troubleshooter engine to pick up the change immediately.
    load_all(force_reload=True)

    result["ai_provider"] = log_entry["ai_provider"]
    return result


def all_correction_feedback() -> List[Dict[str, Any]]:
    """
    Every confirmed (err_msg -> matched KB error pattern) pair logged by
    submit_correction() so far — the ILT Troubleshooter equivalent of
    troubleshooter.psld_semantic_engine.all_feedback(), used by
    troubleshooter/ilt_ai_core.py to train its Level 2 supervised
    MLP+RandomForest ensemble. Only includes entries where the
    correction was actually attached to an existing KB pattern (a real
    confirmed match), never brand-new KB creations (nothing to "confirm"
    there) or "logged_only"/"pending_request" entries.
    """
    rows = []
    for entry in get_corrections_history(limit=10_000):
        err_msg = (entry.get("err_msg") or "").strip()
        matched_pattern = (entry.get("matched_pattern") or "").strip()
        if err_msg and matched_pattern:
            rows.append({"err_msg": err_msg, "matched_pattern": matched_pattern})
    return rows


def log_confirmed_match(err_msg: str, matched_pattern: str, cws: str, source: str = "autonomous_fix") -> None:
    """
    Lightweight logging path for a confirmed (err_msg -> KB pattern)
    association that does NOT go through submit_correction()'s full
    flow — used when troubleshooter.autonomous_fix approves an
    auto-drafted fix: the KB already has the right resolution for this
    pattern (that's exactly why the AI was confident enough to draft the
    fix), so there's no new action/meaning/how-to-check text to merge in
    like a real human correction would bring. All this needs to do is
    feed troubleshooter.ilt_ai_core's Level 2 supervised ensemble one
    more confirmed training example, without appending noisy
    "[Updated ...]" text onto the KB entry's action field every single
    time the same known fix gets reapplied to yet another shipment/date
    variant of the same underlying problem.
    """
    log_entry = {
        "timestamp": datetime.now().isoformat(timespec="seconds"),
        "err_msg": err_msg,
        "matched_pattern": matched_pattern,
        "kb_action": "confirmed_existing",
        "ai_provider": source,
        "cws": cws,
    }
    with json_transaction(CORRECTIONS_LOG_PATH, default=[]) as entries:
        entries.append(log_entry)
        del entries[:-1000]
