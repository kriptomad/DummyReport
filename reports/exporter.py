import pandas as pd
import io

# Leading characters that Excel/Sheets/Calc will interpret as the start of
# a formula if left unescaped in a cell. Also guards against DDE payloads
# starting with a tab/carriage-return followed by one of these chars.
_FORMULA_TRIGGER_CHARS = ("=", "+", "-", "@", "\t", "\r")


def _escape_formula(value):
    """Prefixes a leading apostrophe onto any string value that would
    otherwise be interpreted as a spreadsheet formula when the exported
    file is opened in Excel/Google Sheets/LibreOffice — this is the
    standard "CSV/Excel formula injection" mitigation (the leading `'`
    forces the cell to be treated as plain text, and is not itself
    displayed by the spreadsheet app)."""
    if isinstance(value, str) and value.startswith(_FORMULA_TRIGGER_CHARS):
        return "'" + value
    return value


def _sanitize_for_spreadsheet(df: pd.DataFrame) -> pd.DataFrame:
    """Returns a copy of `df` with formula-injection-triggering string
    values neutralized. Only string/object columns are touched; numeric/
    date columns are returned unchanged. Applied to every CSV/Excel export
    since export contents frequently include free-text fields a user
    typed (KB corrections, request messages, shipment notes, etc.) that
    could otherwise carry a malicious formula payload."""
    sanitized = df.copy()
    for col in sanitized.columns:
        if sanitized[col].dtype == object:
            sanitized[col] = sanitized[col].map(_escape_formula)
    return sanitized


def df_to_csv_bytes(df: pd.DataFrame) -> bytes:
    """Returns a DataFrame as UTF-8 CSV bytes for st.download_button."""
    return _sanitize_for_spreadsheet(df).to_csv(index=False).encode("utf-8")


def df_to_excel_bytes(df: pd.DataFrame, sheet_name: str = "Report") -> bytes:
    """
    Returns a DataFrame as a nicely formatted Excel (.xlsx) for
    st.download_button: bold/colored header row, auto-width columns,
    frozen header row, and an autofilter — a consistent, predefined
    layout used by every export button across the app (Report, Batch,
    Troubleshooter, Tariff, etc.).
    """
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    df = _sanitize_for_spreadsheet(df)

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.sheets[sheet_name]

        header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col_idx, col_cells in enumerate(ws.columns, start=1):
            header_cell = col_cells[0]
            header_cell.fill = header_fill
            header_cell.font = header_font
            header_cell.alignment = Alignment(horizontal="center", vertical="center")

            max_len = max(
                (len(str(cell.value)) if cell.value is not None else 0)
                for cell in col_cells
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 60)

        if ws.max_row > 1 and ws.max_column > 0:
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

    return buffer.getvalue()

