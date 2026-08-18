"""
reports/troubleshoot_report.py
================================
One-click, shareable "Troubleshooting Report" export.

Why this exists
----------------
Before this, the only way to hand off a troubleshooting analysis to a
colleague/manager was screenshots or re-running the same audit query
yourself. This module bundles everything the Troubleshooter tab already
computed — the raw audit rows, the matched KB explanation (meaning / how
to validate / recommended action), the tailored step-by-step checklist,
KB freshness/ownership, and any errors that had NO match in the KB (a
worklist of gaps to feed back into the KB) — into a single, nicely
formatted multi-sheet Excel workbook.

Sheets produced:
  1. "Summary"          — headline counters + error category breakdown.
  2. "Error Analysis"    — one row per unique error: category, KB match
                           score, meaning, how-to-validate, recommended
                           action, suggested owner, freshness, KB
                           owner/last-updated.
  3. "Step-by-Step"      — one row per (error, step) so the checklist is
                           readable/printable outside the app.
  4. "Unmatched Errors"  — errors that had NO KB match at all — useful as
                           a worklist for "these need a new KB fix".
  5. "Raw Audit Data"    — the underlying DEMO_AUDIT rows, for reference.
"""
from __future__ import annotations

import io
from datetime import datetime
from typing import Any, Dict, List

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from troubleshooter.loader import (
    COL_ERROR_PATTERN,
    COL_MEANING,
    COL_HOW_TO_CHECK,
    COL_ACTION,
    COL_RESPONSIBLE,
    is_success_message,
)
from troubleshooter import kb_ownership

HEADER_FILL = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")

# Freshness label kept simple/ASCII for Excel (no HTML/emoji styling needed
# here — a plain text column is enough for a printable report).
_FRESHNESS_LABEL = {"green": "Recent (< 3 months)", "yellow": "Aging (3-12 months)", "red": "Stale (1y+) - review"}


def _style_sheet(ws) -> None:
    """Applies the same header/column-width/freeze/filter style used across
    the app's other Excel exports (see reports/exporter.py)."""
    if ws.max_row < 1 or ws.max_column < 1:
        return
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        header_cell = col_cells[0]
        header_cell.fill = HEADER_FILL
        header_cell.font = HEADER_FONT
        header_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 70)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def build_troubleshoot_report_bytes(
    df_audit: pd.DataFrame,
    error_results: List[Dict[str, Any]],
    shipment_label: str = "",
) -> bytes:
    """
    Builds the full multi-sheet Troubleshooting Report workbook.

    Args:
        df_audit: the raw DEMO_AUDIT rows being analyzed (same DataFrame
            passed to render_troubleshoot_results()).
        error_results: the list of dicts returned by
            troubleshooter.engine.match_errors() — already computed by the
            Troubleshooter tab, so this function does no extra KB matching.
        shipment_label: optional free-text label (e.g. a Shipment ID or
            batch name) shown in the Summary sheet.

    Returns:
        bytes of the .xlsx workbook, ready for st.download_button.
    """
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # ── Sheet 2 data: Error Analysis (one row per unique error) ──────
    analysis_rows = []
    step_rows = []
    unmatched_rows = []

    for item in error_results:
        err_msg = str(item.get("err_msg", "")).strip()
        if not err_msg or err_msg.lower() == "nan":
            continue

        matches = item.get("matches") or []
        category = item.get("category", "")
        steps = item.get("steps") or []

        if not matches:
            unmatched_rows.append({
                "Error Message": err_msg,
                "Classified Category": category or "(unclassified)",
                "Suggested Next Step": (steps[0] if steps else "No specific guidance available — consider adding a new KB fix for this error."),
            })
            continue

        top = matches[0]
        score_pct = round(float(top.get("_match_score", 0.0)) * 100, 1)
        pattern = top.get(COL_ERROR_PATTERN, err_msg)
        kb_meta = kb_ownership.get_meta(pattern)
        fresh_color, fresh_detail = kb_ownership.freshness(kb_meta.get("updated_at", ""))

        analysis_rows.append({
            "Error Message":        err_msg,
            "Category":             category,
            "Match Confidence (%)": score_pct,
            "Probable Meaning":     top.get(COL_MEANING, "—"),
            "How to Validate":      top.get(COL_HOW_TO_CHECK, "—"),
            "Recommended Action":   top.get(COL_ACTION, "—"),
            "Suggested Owner":      top.get(COL_RESPONSIBLE, "—"),
            "Needs Rate Card Lookup Query": "Yes" if item.get("needs_tariff") else "No",
            "KB Freshness":         _FRESHNESS_LABEL.get(fresh_color, fresh_color),
            "Freshness Detail":     fresh_detail,
            "KB Created By":        kb_meta.get("created_by", "SYSTEM"),
            "KB Created At":        str(kb_meta.get("created_at", ""))[:19],
            "KB Last Updated By":   kb_meta.get("updated_by", "SYSTEM"),
            "KB Last Updated At":   str(kb_meta.get("updated_at", ""))[:19],
        })

        for i, step in enumerate(steps, start=1):
            step_rows.append({
                "Error Message": err_msg,
                "Category":      category,
                "Step #":        i,
                "Step Detail":   step,
            })

    df_analysis  = pd.DataFrame(analysis_rows)
    df_steps     = pd.DataFrame(step_rows)
    df_unmatched = pd.DataFrame(unmatched_rows)

    # ── Sheet 1 data: Summary ──────────────────────────────────────
    total_records = len(df_audit)
    unique_shipments = df_audit["SHIPMENT_ID"].nunique() if "SHIPMENT_ID" in df_audit.columns else None
    if "ERR_MSG" in df_audit.columns:
        non_null_errs = df_audit["ERR_MSG"].dropna()
        unique_errors = int(non_null_errs[~non_null_errs.apply(is_success_message)].nunique())
    else:
        unique_errors = len(error_results)

    summary_rows = [
        {"Metric": "Report generated at", "Value": generated_at},
        {"Metric": "Scope", "Value": shipment_label or "(all analyzed rows)"},
        {"Metric": "Total audit records", "Value": total_records},
        {"Metric": "Unique shipments", "Value": unique_shipments if unique_shipments is not None else "—"},
        {"Metric": "Unique error messages", "Value": unique_errors},
        {"Metric": "Errors matched to a KB fix", "Value": len(analysis_rows)},
        {"Metric": "Errors with NO KB match (gaps)", "Value": len(unmatched_rows)},
    ]
    df_summary = pd.DataFrame(summary_rows)

    if not df_analysis.empty:
        cat_counts = (
            df_analysis["Category"].replace("", "(unclassified)").value_counts().reset_index()
        )
        cat_counts.columns = ["Category", "Count"]
    else:
        cat_counts = pd.DataFrame(columns=["Category", "Count"])

    # ── Write workbook ──────────────────────────────────────────────
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df_summary.to_excel(writer, index=False, sheet_name="Summary", startrow=0)
        # Category breakdown appended a couple rows below the headline
        # metrics table, same sheet, so it's a single "at a glance" tab.
        cat_start_row = len(df_summary) + 3
        cat_counts.to_excel(writer, index=False, sheet_name="Summary", startrow=cat_start_row)

        if not df_analysis.empty:
            df_analysis.to_excel(writer, index=False, sheet_name="Error Analysis")
        if not df_steps.empty:
            df_steps.to_excel(writer, index=False, sheet_name="Step-by-Step")
        if not df_unmatched.empty:
            df_unmatched.to_excel(writer, index=False, sheet_name="Unmatched Errors")

        # Raw audit data for full traceability — capped defensively so a
        # very large batch export doesn't produce an unwieldy workbook.
        df_raw = df_audit.head(20000)
        df_raw.to_excel(writer, index=False, sheet_name="Raw Audit Data")

        for sheet_name in writer.sheets:
            _style_sheet(writer.sheets[sheet_name])

    return buffer.getvalue()
